from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font
from openpyxl.cell import Cell
import configparser
import os
import glob
from NotHighlighted import FoundTagOwner


class ExcelOperations:

    docOwner = ""

    def __init__(self, configObject, docPath):
        self.config_object = configObject
        print("loading excel workbook...")
        self.excel_config = self.config_object["ExcelConfig"]
        #self.wb = load_workbook(filename=docPath + "/" + self.excel_config["WorkBookName"])
        # self.ws = self.wb[self.excel_config["SheetName"]]
        # self.row_count = self.ws.max_row
        # self.doc_num_cell = self.excel_config["DocNumCell"]
        # self.tag_num_cell = self.excel_config["TagNumCell"]

    def CreateDirectory(self, dirPath):
        if not os.path.exists(dirPath):
            os.mkdir(dirPath)

    def readTagList(self, excelPath):
        tagLists = []
        print("loading excel workbook...")
        excel_config = self.config_object["ExcelConfig"]
        wb = load_workbook(filename=excelPath + "\\" +
                           excel_config["WorkBookName"])
        ws = wb[excel_config["SheetName"]]
        row_count = ws.max_row
        tag_num_cell = excel_config["TagNumCell"]
        # getting doc number
        for cell in ws[tag_num_cell + excel_config["StartRowIndex"] + ":" + tag_num_cell + str(row_count)]:
            # getting tag number
            tagLists.append(ws[tag_num_cell + str(cell[0].row)].value)
        return tagLists

    def getUnknownTagList(self, fPath, pdfFiles):
        docTag = {}
        print("loading excel workbook...")
        excel_config = self.config_object["ExcelConfig"]
        wb = load_workbook(filename=fPath + "\\" +
                           excel_config["WorkBookName"])
        ws = wb[excel_config["SheetName"]]
        row_count = ws.max_row
        doc_num_cell = excel_config["DocNumCell"]
        tag_num_cell = excel_config["TagNumCell"]
        for pdf in pdfFiles:
            tagLists = []
            pdfName = os.path.basename(pdf)  # getting filename from file path
            # getting doc number
            for cell in ws[doc_num_cell + excel_config["StartRowIndex"] + ":" + doc_num_cell + str(row_count)]:
                if cell[0].value + "_R" + ws[excel_config["DocRevisionCell"] + str(cell[0].row)].value + ".pdf" == pdfName:
                    # getting tag number
                    tagLists.append(ws[tag_num_cell + str(cell[0].row)].value)
            if pdfName not in docTag:
                docTag[pdfName] = tagLists
            elif pdfName in docTag:
                docTag[pdfName] = docTag[docTag[pdfName], tagLists]
        return docTag

    def createOutputExcelFile(self, fpath, notFoundDict, foundDict, foundTagRow):
        excelFile = glob.glob(
            "Output\\" + foundTagRow[0].docOwner + "\\Output_UnknownTag.xlsx")
        wb = Workbook()
        row_index = 2
        self.CreateDirectory("Output\\"+foundTagRow[0].docOwner)
        if not excelFile:
            wsnf = wb.create_sheet("NotFound_UnknownTag")
            wsf = wb.create_sheet("Found_UnknownTag")
            self.setHeaders(wsnf)
            self.setHeaders(wsf)
            for key, value in notFoundDict.items():
                for val in value:
                    wsnf["A" + str(row_index)].value = key
                    wsnf["B" + str(row_index)].value = val.docRevision
                    wsnf["C" + str(row_index)].value = val.docTitle
                    wsnf["D" + str(row_index)].value = val.tagNumber
                    wsnf["E" + str(row_index)].value = val.docType
                    row_index += 1
            row_index = 2
            for key, value in foundDict.items():
                for val in value:
                    wsf["A" + str(row_index)].value = key
                    wsf["B" + str(row_index)].value = val.docRevision
                    wsf["C" + str(row_index)].value = val.docTitle
                    wsf["D" + str(row_index)].value = val.tagNumber
                    wsf["E" + str(row_index)].value = val.docType
                    row_index += 1
            if wb.get_sheet_by_name("Sheet"):
                wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
            wb.save("Output\\" +
                    foundTagRow[0].docOwner + "\\Output_UnknownTag.xlsx")
        else:
            wb = load_workbook(
                filename="Output\\" + foundTagRow[0].docOwner + "\\Output_UnknownTag.xlsx")
            wsnf = wb["NotFound_UnknownTag"]
            wsf = wb["Found_UnknownTag"]
            row_index = wsnf.max_row + 1
            self.setHeaders(wsnf)
            self.setHeaders(wsf)
            for key, value in notFoundDict.items():
                for val in value:
                    wsnf["A" + str(row_index)].value = key
                    wsnf["B" + str(row_index)].value = val.docRevision
                    wsnf["C" + str(row_index)].value = val.docTitle
                    wsnf["D" + str(row_index)].value = val.tagNumber
                    wsnf["E" + str(row_index)].value = val.docType
                    row_index += 1
            row_index = wsf.max_row + 1
            for key, value in foundDict.items():
                for val in value:
                    wsf["A" + str(row_index)].value = key
                    wsf["B" + str(row_index)].value = val.docRevision
                    wsf["C" + str(row_index)].value = val.docTitle
                    wsf["D" + str(row_index)].value = val.tagNumber
                    wsf["E" + str(row_index)].value = val.docType
                    row_index += 1
            wb.save("Output\\" +
                    foundTagRow[0].docOwner + "\\Output_UnknownTag.xlsx")

    def setHeaders(self, worksheet):
        worksheet["A1"] = "Document Number"
        worksheet["B1"] = "Document Revision"
        worksheet["C1"] = "Document Title"
        worksheet["D1"] = "Tag Number"
        worksheet["E1"] = "Doc Type"

    def designCell(self, docCell: Cell, tagCell: Cell):
        docCell.style.font.bold = True
        docCell.fill = PatternFill(bgColor="FFFFFF00", fill_type="solid")
        tagCell.style.font.bold = True
        tagCell.fill = PatternFill(bgColor="FFFFFF00", fill_type="solid")

    def createDocTag(self, doc_tag_dict):
        excelFile = glob.glob("Output\\Output_doc_tag.xlsx")
        wb = Workbook()
        row_index = 2
        self.CreateDirectory("Output")
        if not excelFile:
            wsdoctag = wb.create_sheet("Doc_Tag")
            self.setHeaders(wsdoctag)
            for key, value in doc_tag_dict.items():
                for tagNumber in value:
                    wsdoctag["A" + str(row_index)].value = key
                    wsdoctag["B" + str(row_index)].value = "val.docRevision"
                    wsdoctag["C" + str(row_index)].value = "val.docTitle"
                    wsdoctag["D" + str(row_index)].value = tagNumber
                    wsdoctag["E" + str(row_index)].value = "val.docType"
                    row_index += 1
            row_index = 2
            if wb.get_sheet_by_name("Sheet"):
                wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
            wb.save("Output\\Output_doc_tag.xlsx")
